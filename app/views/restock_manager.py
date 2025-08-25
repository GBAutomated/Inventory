import streamlit as st
import pandas as pd
import tempfile, io, os
import numpy as np
import altair as alt
from io import BytesIO
from datetime import date
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from app.services.supabase_uploader import (
    insert_restock_qt,
    get_item_by_name,
    fetch_restock_kpi_source
)

STATUS_ORDER = ["Critical", "Reorder now", "Near", "Healthy"]
URGENCY_PRIORITY = {"Critical": 0, "Reorder now": 1, "Near": 2, "Healthy": 3}

def show_restock_form():
    st.subheader("ğŸ› ï¸ Restock Items")

def generate_restock_file_by_categories_template(items: list) -> bytes:
    df = pd.DataFrame(items)
    rename_map = {"category_name": "Category", "name": "Name", "description": "Description"}
    df = df.rename(columns=rename_map)

    df = df[["Category", "Name", "Description"]]
    df["Reorder Qty"] = ""   # celdas vacÃ­as para que el usuario diligencie

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = "ReorderingQuantities"
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)
        sheet = writer.sheets[sheet_name]

        sheet["A1"] = "Reordering Minimun Quantities"
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

        sheet.freeze_panes = "A3"

        col_widths = [18, 22, 50, 14]
        for i, w in enumerate(col_widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = w

        indent_align = Alignment(indent=1, vertical="top", wrap_text=True)
        right_align  = Alignment(indent=1, vertical="top")

        for cell in sheet[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")

        thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))

        first_row = 2      
        last_row  = len(df) + 2 
        first_col, last_col = 1, 4

        for row in sheet.iter_rows(min_row=first_row, max_row=last_row,
                                   min_col=first_col, max_col=last_col):
            for j, cell in enumerate(row, start=1):
                cell.border = thin
                if j in (1, 2, 3):
                    cell.alignment = indent_align
                else:
                    cell.alignment = right_align

    output.seek(0)
    return output.getvalue()

def show_upload_restock_file(user: int):    
    st.subheader("ğŸ“¤ Upload Restock File")
    uploaded_file = st.file_uploader(
        "Select the file with the Restock min quantities",
        type=["xlsx"]
    )
    today = datetime.today().isoformat()

    if not uploaded_file:
        return

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.read())
        temp_path = tmp.name

    try:
        df = pd.read_excel(temp_path, header=None)
        st.dataframe(df.head(6).astype("string"))

        raw_headers = df.iloc[1].tolist()
        headers = [str(h).strip().lower() for h in raw_headers]
        expected_headers = ["category", "name", "description", "reorder qty"]

        if headers != expected_headers:
            st.error(f"âŒ Header mismatch.\nExpected: {expected_headers}\nFound: {headers}")
            return

        df_data = df.iloc[2:].copy()
        df_data.columns = expected_headers

        df_data = df_data.astype("string")
        for c in df_data.columns:
            df_data[c] = (df_data[c]
                          .str.replace("\u00a0", " ", regex=False)
                          .str.strip())

        df_data = df_data.dropna(subset=["name"])
        df_data["reorder qty"] = (df_data["reorder qty"]
                                  .str.replace(r"[^\d\.\-]", "", regex=True)
                                  .replace({"": None, "nan": None}))
        df_data["reorder qty"] = pd.to_numeric(df_data["reorder qty"], errors="coerce").fillna(0)

        st.dataframe(df_data.head().astype("string"))

        restock_items = []
        total_steps = len(df_data) + 2
        progress = st.progress(0)
        status = st.empty()
        step = 0

        for i, (_, row) in enumerate(df_data.iterrows(), 1):
            status.info(f"ğŸ” Matching item {i} of {len(df_data)}: {row['name']}")
            item = get_item_by_name(row["name"])
            if not item:
                st.warning(f"âš ï¸ Item not found: {row['name']}")
                continue

            restock_items.append({
                "id_item": item["id"],
                "id_user": user,
                "date": today,
                "restock_qty": float(row["reorder qty"]),
            })
            step += 1
            progress.progress(step / total_steps)

        status.info(f"ğŸ“¤ Uploading {len(restock_items)} items to Supabase...")
        items_ok = insert_restock_qt(restock_items)
        step += 1
        progress.progress(step / total_steps)

        if items_ok:
            status.success("âœ… Restock quantities successfully uploaded.")
        else:
            status.error("âŒ Failed to upload Restock quantities for items.")
        progress.empty()

    except Exception as e:
        st.error(f"âŒ Error processing file: {str(e)}")
    finally:
        os.unlink(temp_path)

def classify_status(avail: float, minq: float) -> str:
    if pd.isna(minq) or minq <= 0:
        return "Healthy"
    if avail < minq:
        return "Critical"
    if avail == minq:
        return "Reorder now"
    near = minq * 1.20
    if avail <= near:
        return "Near"
    return "Healthy"

def build_kpis(rows: list[dict]) -> tuple[pd.DataFrame, dict]:
    df = pd.DataFrame(rows)
    if df.empty:
        return df, {"Critical": 0, "Reorder now": 0, "Near": 0, "Healthy": 0}

    for c in ["available", "on_so", "on_po", "restock_qty"]:
        df[c] = pd.to_numeric(df.get(c), errors="coerce").fillna(0)

    df["status"] = [
        classify_status(avail, minq)
        for avail, minq in zip(df["available"], df["restock_qty"])
    ]

    kpis = df["status"].value_counts().reindex(
        ["Critical", "Reorder now", "Near", "Healthy"], fill_value=0
    ).to_dict()

    return df, kpis

def show_kpis():
    st.subheader("ğŸ“Š Items Stock KPI's")

    data = fetch_restock_kpi_source()
    df, kpis = build_kpis(data)
    df["status"] = pd.Categorical(df["status"], categories=STATUS_ORDER, ordered=True)

    crit    = int(kpis.get("Critical", 0))
    reorder = int(kpis.get("Reorder now", 0))
    near    = int(kpis.get("Near", 0))
    healthy = int(kpis.get("Healthy", 0))

    def fmt(n: int) -> str:
        return f"{n:,} item" + ("s" if n != 1 else "")

    col1, col2, col3, col4 = st.columns(4)

    col1.metric(
        label="Critical",
        value=fmt(crit),
        delta="- ğŸš¨ Stock below min",
        border=True
    )
    col2.metric(
        label="Reorder Now",
        value=fmt(reorder),
        delta="- ğŸ”´ Needs reorder",
        border=True
    )
    col3.metric(
        label="Near",
        value=fmt(near),
        delta="ğŸŸ  Near min",
        delta_color="off", 
        border=True  
    )
    col4.metric(
        label="Healthy",
        value=fmt(healthy),
        delta="ğŸŸ¢ Healthy",
        border=True
    )

def show_restock_table_and_file_download():

    st.subheader("ğŸ“ Create Restock Items List")
    rows = fetch_restock_kpi_source()
    if not rows:
        st.info("No data to show.")
        return

    df = pd.DataFrame(rows)
    for c in ["available", "on_so", "on_po", "restock_qty"]:
        df[c] = pd.to_numeric(df.get(c, 0), errors="coerce").fillna(0)

    def classify(avail, minq):
        if pd.isna(minq) or minq <= 0:
            return "Healthy"
        if avail < minq:
            return "Critical"
        if avail == minq:
            return "Reorder now"
        if avail <= minq * 1.20:
            return "Near"
        return "Healthy"

    STATUS_ORDER = ["Critical", "Reorder now", "Near", "Healthy"]
    URGENCY_PRIORITY = {"Critical": 0, "Reorder now": 1, "Near": 2, "Healthy": 3}

    df["status"] = [classify(a, m) for a, m in zip(df["available"], df["restock_qty"])]
    df["status"] = pd.Categorical(df["status"], categories=STATUS_ORDER, ordered=True)
    df["difference"] = (df["restock_qty"] - df["available"]).astype(float)
    df["urgency"]    = df["status"].map(URGENCY_PRIORITY)

    df = df.sort_values(["urgency", "difference", "on_so"], ascending=[True, False, False])

    view = pd.DataFrame({
        "item_id":     df["item_id"] if "item_id" in df.columns else df["name"],
        "code":        df["name"] if "name" in df.columns else "",
        "item":        (df["description"] if "description" in df.columns else
                        (df["name"] if "name" in df.columns else "")),
        "category":    (df["category_name"] if "category_name" in df.columns else ""),
        "available":   df["available"],
        "reorder_min": df["restock_qty"],
        "difference":  df["difference"],
        "status":      df["status"].astype(str),
    })

    default_selected = view["status"].isin(["Critical", "Reorder now"])

    EMOJI = {"Critical": "ğŸš¨", "Reorder now": "ğŸ”´", "Near": "ğŸŸ ", "Healthy": "ğŸŸ¢"}
    view["status"] = view["status"].map(lambda s: f"{EMOJI.get(s, s)} {s}")

    index_col = "item_id" if "item_id" in view.columns else "code"
    view = view.set_index(index_col)
    view.insert(0, "select", default_selected.values)

    cols_view = ["select", "code", "item", "category", "available", "reorder_min", "difference", "status"]
    edited = st.data_editor(
        view[cols_view],
        hide_index=True,
        num_rows="fixed",
        use_container_width=True,
        column_config={
            "select":       st.column_config.CheckboxColumn("Pick", help="Add/Remove"),
            "code":         st.column_config.TextColumn("Code", width="small"),
            "item":         st.column_config.TextColumn("Item"),
            "category":     st.column_config.TextColumn("Category", width="small"),
            "available":    st.column_config.NumberColumn("Available", format="%.0f", width="small"),
            "reorder_min":  st.column_config.NumberColumn("Reorder Qty", format="%.0f", width="small"),
            "difference":   st.column_config.NumberColumn("Difference", format="%.0f"),
            "status":       st.column_config.TextColumn("Status", width="small"),
        },
        key="restock_pick_editor",
    )

    selected = edited[edited["select"]].copy()
    st.caption(f"Selected: **{len(selected)}** items")

    if len(selected) == 0:
        st.warning("Pick one or more rows to enable the download.")
        return

    export_df = selected[["code", "item", "category", "available", "reorder_min", "difference"]].reset_index(drop=True)
    export_df = export_df.rename(columns={
        "reorder_min": "Reorder Qty",
        "difference":  "Difference (Reorder - Available)"
    })

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet_name = "RestockItems"

        HEADER_ROW = 3
        DATA_START_ROW = HEADER_ROW + 1
        STARTROW_TO_EXCEL = HEADER_ROW - 1

        export_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=STARTROW_TO_EXCEL)
        ws = writer.sheets[sheet_name]

        last_col = export_df.shape[1]
        ws["A1"] = "Items to Restock"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        from openpyxl.styles import Font, Border, Side, Alignment
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = f"A{DATA_START_ROW}"

        for cell in ws[HEADER_ROW]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))
        last_row = HEADER_ROW + len(export_df)
        for row in ws.iter_rows(min_row=HEADER_ROW, max_row=last_row,
                                min_col=1, max_col=last_col):
            for cell in row:
                cell.border = thin

        from openpyxl.utils import get_column_letter
        widths = [18, 48, 22, 12, 14, 24]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for r in range(DATA_START_ROW, last_row + 1):
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

    buf.seek(0)
    st.download_button(
        "ğŸ“¥ Download Restock File",
        data=buf.getvalue(),
        file_name=f"restock_items_list_{date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
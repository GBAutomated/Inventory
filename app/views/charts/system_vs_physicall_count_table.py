import streamlit as st
import pandas as pd
from app.services.supabase_uploader import fetch_inventory_comparison
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def show_inventory_comparison():
    st.subheader("System VS Physicall Count")

    data = fetch_inventory_comparison()
    if not data:
        st.warning("No inventory comparison data found.")
        return

    df = pd.DataFrame(data)
    df = df[df["difference"] != 0]

    if df.empty:
        st.success("✅ No discrepancies found. System and physical counts match.")
        return

    st.info(f"🔍 {len(df)} items with discrepancies found.")

    display_df = df.rename(columns={
    "count_date": "Count Date",
    "name": "Item Code",
    "description": "Description",
    "category_name": "Category",
    "on_hand": "System Qty",
    "counted_qty": "Counted Qty",
    "difference": "Difference",
    "responsable": "Responsible",
    "notes": "Notes"
})

    st.dataframe(display_df[["Item Code", "Description", "System Qty", "Counted Qty", "Difference", "Notes"]])

    excel_data = to_excel_bytes(df)
    st.download_button(
    label="📥 Download Excel Report",
    data=excel_data,
    file_name="inventory_comparison.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    df = df.rename(columns={
        "count_date": "Count Date",
        "name": "Item Code",
        "description": "Description",
        "category": "Category",
        "on_hand": "System Qty",
        "counted_qty": "Counted Qty",
        "difference": "Difference",
        "responsable": "Responsible",
        "notes": "Notes"
    })

    if "item_id" in df.columns:
        df = df.drop(columns=["item_id"])


    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Comparison"

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_idx, column in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = max_length + 4
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    wb.save(output)
    output.seek(0)
    return output.read()